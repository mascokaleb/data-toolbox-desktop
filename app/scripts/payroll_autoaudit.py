"""
name: Payroll auto‑audit
description: Runs automated checks on the payroll workbook and produces a flagged‑rows sheet.
required_files:
  payroll_xlsx: "Payroll export (XLSX)"
outputs:
  audit_xlsx: "Audit results (XLSX)"
"""
from pathlib import Path
import pandas as pd
import numpy as np
from dateutil.relativedelta import relativedelta
from openpyxl.styles import PatternFill
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

pd.options.mode.chained_assignment = None

# ---------- helpers ----------------------------------------------------------
def first_of_month_following(d):
    return (d + relativedelta(day=1, months=+1)).normalize()

def flag(cond, label, errors, highlight_dict, cols_to_highlight=None):
    """Append label to Errors; record failing row indices for the given columns."""
    errors.loc[cond, "Errors"] += f"{label}; "
    if cols_to_highlight:
        failing_idx = errors[cond].index
        for col in cols_to_highlight:
            highlight_dict[col].update(failing_idx)

# ---------- load -------------------------------------------------------------
def main(payroll_xlsx: Path) -> Path:
    src = payroll_xlsx
    dst = payroll_xlsx.with_name("payroll_auto_audit_results.xlsx")
    df = pd.read_excel(src)

    # make an empty error column
    df["Errors"] = ""

    # track which cells need a red fill
    highlight_dict = {col: set() for col in df.columns}

    # ---------- data hygiene -----------------------------------------------------
    # Strip leading/trailing whitespace but only on cells that are actually strings
    for col in df.columns:
        if df[col].dtype == "object":
            df[col] = df[col].apply(lambda x: x.strip() if isinstance(x, str) else x)

    # ---------- reference tables -------------------------------------------------
    # All valid Department/Location combinations
    dept_loc_ok = {
        ("Legal", "Atlanta"), ("Legal", "Baltimore"), ("Legal", "Boston"), ("Legal", "Charlotte"),
        ("Legal", "Chicago"), ("Legal", "Cleveland"), ("Legal", "Columbia"), ("Legal", "Columbus"),
        ("Legal", "Dallas"), ("Legal", "Denver"), ("Legal", "Detroit"), ("Legal", "Fort Lauderdale"),
        ("Legal", "Gulfport"), ("Legal", "Houston"), ("Legal", "Irvine"), ("Legal", "Kansas City"),
        ("Legal", "Las Vegas"), ("Legal", "Los Angeles"), ("Legal", "Louisville"), ("Legal", "McLean"),
        ("Legal", "Memphis"), ("Legal", "Nashville"), ("Legal", "New Jersey"), ("Legal", "New Orleans"),
        ("Legal", "New York"), ("Legal", "Orlando"), ("Legal", "Philadelphia"), ("Legal", "Phoenix"),
        ("Legal", "Pittsburgh"), ("Legal", "Portland"), ("Legal", "Portland ME"), ("Legal", "Sacramento"),
        ("Legal", "San Diego"), ("Legal", "San Francisco"), ("Legal", "Seattle"), ("Legal", "Tampa"),
        ("Legal", "Washington DC"), ("Legal", "Woodland Hills"),
        ("Legal", "Virtual Office"),
        ("Legal", "Birmingham"),
        ("Legal Operations", "National Office"), ("Legal Operations", "Virtual Office"),
        ("Accounting", "Corporate"), ("Administration", "Corporate"), ("Attorney Recruiting", "Corporate"),
        ("Billing and Collections", "Corporate"), ("Content", "Corporate"), ("Facilities", "Corporate"),
        ("Financial Analysis", "Corporate"), ("Human Resources", "Corporate"),
        ("Information Governance", "Corporate"), ("Information Security", "Corporate"),
        ("Information Technology", "Corporate"), ("Knowledge Management", "Corporate"),
        ("Legal", "Corporate"), ("Legal Operations", "Corporate"), ("Library", "Corporate"),
        ("Marketing", "Corporate"), ("New Business Intake", "Corporate"), ("Partner Recruiting", "Corporate"),
        ("Professional Development", "Corporate"),
    }

    # Simple pairings
    payfreq_group_ok = {("B", "B"), ("M", "M"), ("Q", "Q"), ("S", "S")}
    group_type_ok    = {("B", "Hourly"), ("M", "Salary"), ("S", "Salary")}

    # Pay‑group specific expected numeric constants
    default_hours_map  = {"B": 80, "M": 173.34, "S": 86.67}
    salary_factor      = {"M": 12, "S": 24}

    # -----------------------------------------------------------------------------


    # ---------- rule checks ------------------------------------------------------

    # 1. Department/Location pair must be allowed
    dept_needing_check = df["Department"].isin({d for d, _ in dept_loc_ok})
    bad = dept_needing_check & ~df[["Department", "Location"]].apply(tuple, axis=1).isin(dept_loc_ok)
    flag(bad, "Dept/Loc invalid", df, highlight_dict, ["Department", "Location"])

    # 1a. Virtual Office → Department depends on Position
    cond_paralegal_staff = (
        df["Location"].eq("Virtual Office")
        & df["Position"].isin(["Paralegal", "Staff"])
        & df["Department"].ne("Legal")
    )
    cond_attorney = (
        df["Location"].eq("Virtual Office")
        & df["Position"].eq("Attorney")
        & df["Department"].ne("Legal Operations")
    )
    bad = cond_paralegal_staff | cond_attorney
    flag(bad, "Virtual Office Dept mismatch", df, highlight_dict, ["Department", "Location", "Position"])

    # 2. Pay Frequency Code ↔ Pay Group
    pf_needing_check = df["Pay Frequency Code"].isin({pf for pf, _ in payfreq_group_ok})
    bad = pf_needing_check & ~df[["Pay Frequency Code", "Pay Group"]].apply(tuple, axis=1).isin(payfreq_group_ok)
    flag(bad, "PayFreq/Group mismatch", df, highlight_dict, ["Pay Frequency Code", "Pay Group"])

    # 3. Pay Group ↔ Pay Type
    pg_needing_check = df["Pay Group"].isin({pg for pg, _ in group_type_ok})
    bad = pg_needing_check & ~df[["Pay Group", "Pay Type"]].apply(tuple, axis=1).isin(group_type_ok)
    flag(bad, "PayGroup/Type mismatch", df, highlight_dict, ["Pay Group", "Pay Type"])

    # 4. Position‑specific blanks

    # Attorney must have a JD year, except when the Job Title is Summer Clerk or Law Clerk
    mask_atty   = df["Position"].str.casefold().eq("attorney")
    mask_clerks = df["Job Title (Point in Time)"].str.casefold().isin(["summer clerk", "law clerk"])
    cond_missing_jd = mask_atty & ~mask_clerks & df["JD Graduation Year"].isna()

    flag(
        cond_missing_jd,
        "Attorney missing JD Graduation Year",
        df,
        highlight_dict,
        ["JD Graduation Year", "Position", "Job Title (Point in Time)"]
    )

    flag(
        (df["Job Title (Point in Time)"].str.casefold().eq("associate") & df["Credited Year (Attorneys Only)"].isna()),
        "Associate missing Credited Year",
        df,
        highlight_dict,
        ["Credited Year (Attorneys Only)", "Job Title (Point in Time)"]
    )

    # 5. Partner dates
    flag(
        (df["Job Title (Point in Time)"].str.casefold().eq("partner income") & df["Income Partner Date"].isna()),
        "Partner Income missing Income Partner Date",
        df,
        highlight_dict,
        ["Job Title (Point in Time)", "Income Partner Date"]
    )

    flag(
        (df["Job Title (Point in Time)"].str.casefold().eq("partner") & df["Equity Partner Date"].isna()),
        "Partner missing Equity Partner Date",
        df,
        highlight_dict,
        ["Job Title (Point in Time)", "Equity Partner Date"]
    )

    # 6. Pay Type ↔ Auto Pay (only check rows where Pay Type is Hourly or Salary)
    hourly_mask  = df["Pay Type"].str.casefold().eq("hourly")
    salary_mask  = df["Pay Type"].str.casefold().eq("salary")


    bad = hourly_mask & df["Auto Pay"].ne("No")
    flag(bad, "Hourly AutoPay≠No", df, highlight_dict, ["Pay Type", "Auto Pay"])

    bad = salary_mask & df["Auto Pay"].ne("Yes")
    flag(bad, "Salary AutoPay≠Yes", df, highlight_dict, ["Pay Type", "Auto Pay"])

    # 6c. Hourly Attorneys must be Pay Group S and Auto Pay = No
    mask_hourly_atty = (
        df["Pay Type"].str.casefold().eq("hourly")
        & df["Position"].str.casefold().eq("attorney")
    )

    bad_pg = mask_hourly_atty & df["Pay Group"].ne("S")
    flag(bad_pg, "Hourly Attorney pay group≠S", df,
         highlight_dict, ["Pay Type", "Position", "Pay Group"])

    bad_ap = mask_hourly_atty & df["Auto Pay"].ne("No")
    flag(bad_ap, "Hourly Attorney AutoPay≠No", df,
         highlight_dict, ["Pay Type", "Position", "Auto Pay"])

    # 7. Aderant Number ↔ Alight Username
    # Expected pattern: FP + zero‑padded 4‑digit Aderant Number
    aderant_raw = df["Aderant Number"].astype(str).str.strip()

    # Drop trailing ".0" if the number came in as a float, keep only digits
    aderant_digits = aderant_raw.str.extract(r"(\d+)")[0]

    expected_username = "FP" + aderant_digits.str.zfill(4)

    aderant_present = aderant_digits.notna() & aderant_digits.ne("")
    bad = aderant_present & df["Alight Username"].fillna("").str.strip().ne(expected_username)

    flag(bad, "Alight Username mismatch", df, highlight_dict, ["Aderant Number", "Alight Username"])

    # 8. Job Title ↔ Security Group (case‑insensitive, and highlight mismatches)
    job_title_cf = df["Job Title (Point in Time)"].str.casefold().fillna("")
    sec_grp_cf   = df["Security Group"].str.casefold().fillna("")

    partner_mask         = job_title_cf.eq("partner")
    partner_mismatch     = partner_mask & ~sec_grp_cf.eq("equity employee")

    nonpartner_mismatch  = (~partner_mask) & ~sec_grp_cf.eq("employee")

    bad = partner_mismatch | nonpartner_mismatch
    flag(bad, "Security Group mismatch", df,
         highlight_dict, ["Job Title (Point in Time)", "Security Group"])



    # 9. Fields that must be non‑blank / unique
    blank_ft = df["FT Hours (37.5 or 40)"].isna()
    flag(blank_ft, "FT Hours blank", df, highlight_dict, ["FT Hours (37.5 or 40)"])

    # Default Hours must be non-blank and non-zero
    blank_def_hours = df["Default Hours"].isna() | (df["Default Hours"] == 0)
    flag(blank_def_hours, "Default Hours blank/zero", df, highlight_dict, ["Default Hours"])


    # Regular Full Time employees must have Default Hours > 59
    mask_rft = df["Employment Type Description"].str.casefold().eq("regular full time")
    bad_rft_hours = mask_rft & (df["Default Hours"].fillna(0) <= 59)
    flag(
        bad_rft_hours,
        "Default Hours too low for Regular Full Time",
        df,
        highlight_dict,
        ["Employment Type Description", "Default Hours"]
    )

    # Regular Full Time or Regular Part Time must have Retirement Plan = FP0102
    mask_emp_type = df["Employment Type Description"].str.casefold().isin(
        ["regular full time", "regular part time"]
    )
    bad_ret_plan = mask_emp_type & df["Retirement Plan"].ne("FP0102")

    flag(
        bad_ret_plan,
        "Retirement Plan mismatch for Reg FT/PT",
        df,
        highlight_dict,
        ["Employment Type Description", "Retirement Plan"]
    )

    blank_cell_dir = df["Do not post cell phone on directory"].isna() | \
                     (df["Do not post cell phone on directory"].astype(str).str.strip() == "")
    flag(blank_cell_dir, "Cell dir flag blank", df, highlight_dict, ["Do not post cell phone on directory"])

    blank_aderant = df["Aderant Number"].isna() | \
                    (df["Aderant Number"].astype(str).str.strip() == "")
    flag(blank_aderant, "Aderant blank", df, highlight_dict, ["Aderant Number"])

    blank_email = df["Current Work Email"].isna() | \
                  (df["Current Work Email"].astype(str).str.strip() == "")
    flag(blank_email, "Email blank", df, highlight_dict, ["Current Work Email"])

    # Duplicate Aderant but only among non‑blank values
    dup_aderant = df["Aderant Number"].notna() & \
                  (df["Aderant Number"].astype(str).str.strip() != "") & \
                  df["Aderant Number"].duplicated(keep=False)
    flag(dup_aderant, "Aderant duplicate", df, highlight_dict, ["Aderant Number"])


    # 10. Pay Group‑specific FTE checks
    fte_col = "FTE"

    # Expected FTE for each Pay Group
    expected_fte = pd.Series(index=df.index, dtype=float)

    # B-group: Default Hours / (FT Hours × 2)
    mask_b = df["Pay Group"].eq("B")
    expected_fte.loc[mask_b] = df.loc[mask_b, "Default Hours"] / (
        df.loc[mask_b, "FT Hours (37.5 or 40)"] * 2
    )

    # M‑group: Default Hours / 173.34
    mask_m = df["Pay Group"].eq("M")
    expected_fte.loc[mask_m] = df.loc[mask_m, "Default Hours"] / 173.34

    # S‑group: Default Hours / 86.67
    mask_s = df["Pay Group"].eq("S")
    expected_fte.loc[mask_s] = df.loc[mask_s, "Default Hours"] / 86.67

    # Compare with tolerance
    fte_present = df[fte_col].notna()
    # Only run this rule for Pay Groups B, M, S
    valid_pg_mask = df["Pay Group"].isin(["B", "M", "S"])
    # allow a small absolute tolerance of ±0.002 (≈0.2 on a 100‑point scale)
    bad = fte_present & valid_pg_mask & ~np.isclose(df[fte_col], expected_fte, atol=0.002, rtol=0)

    flag(bad, "FTE mismatch", df, highlight_dict, [fte_col, "Default Hours", "Pay Group"])



    # 11. Profit-sharing Eligibility Date = 1st day of the month 13 months after Hire Date
    hire_present = df["Hire Date"].notna()

    # Normalize to midnight so time components don’t break equality
    hire_norm = pd.to_datetime(df["Hire Date"]).dt.normalize()
    pse_norm  = pd.to_datetime(df["Profit Sharing Eligibility Date"]).dt.normalize()

    # If hired on the 1st of a month → exactly one year later; otherwise → first day 13 months later
    expected_pse = hire_norm.apply(
        lambda d: (d + relativedelta(years=1)).normalize() if d.day == 1
        else (d + relativedelta(months=+13, day=1)).normalize()
    )


    # Enforce rule only when Hire Date is between 2016‑01‑01 and 2025‑06‑01
    # and any Rehire Date (if present) is also ≤ 2025‑06‑01
    rehire_norm = pd.to_datetime(df["Rehire Date"]).dt.normalize()
    mask_recent_hire = (
        (hire_norm >= pd.Timestamp("2016-01-01")) &
        (hire_norm <= pd.Timestamp("2025-06-01")) &
        (rehire_norm.isna() | (rehire_norm <= pd.Timestamp("2025-06-01")))
    )

    missing_pse  = mask_recent_hire & hire_present & pse_norm.isna()
    mismatch_pse = mask_recent_hire & hire_present & pse_norm.notna() & (pse_norm != expected_pse)

    bad = missing_pse | mismatch_pse

    flag(
        bad,
        "Profit-share date wrong",
        df,
        highlight_dict,
        ["Profit Sharing Eligibility Date", "Hire Date"]
    )


    # 12. Annual ↔ Per‑check Salary (only for Pay Groups M and S)
    salary_factor_map = {"M": 12, "S": 24}  # divisor = pay periods per year
    factor_series = df["Pay Group"].map(salary_factor_map)

    valid_pg_mask = df["Pay Group"].isin(["M", "S"])
    expected_per_check = df["Annual Salary"] / factor_series

    # allow small rounding differences (±$0.01)
    bad = valid_pg_mask & ~np.isclose(df["Per Check Salary"], expected_per_check, atol=0.01, rtol=0)

    flag(
        bad,
        "Per‑check Salary mismatch",
        df,
        highlight_dict,
        ["Per Check Salary", "Annual Salary", "Pay Group"]
    )

    # 13. AS FTE % validation
    asfte_col = "AS FTE % (/37.5)"
    expected_asfte = pd.Series(index=df.index, dtype=float)

    # Virtual Office Attorneys 
    mask_vo_attorney = (
        df["Location"].eq("Virtual Office")
        & df["Position"].str.casefold().eq("attorney")
    )
    expected_asfte.loc[mask_vo_attorney] = (
        df.loc[mask_vo_attorney, "Default Hours"] * 24 / 1850 * 100
    )

    # Pay‑group based calculations 
    mask_b = df["Pay Group"].eq("B")   # Bi‑weekly
    mask_s = df["Pay Group"].eq("S")   # Semi‑monthly
    mask_m = df["Pay Group"].eq("M")   # Monthly

    expected_asfte.loc[mask_b] = df.loc[mask_b, "Default Hours"] * 26 / 1950 * 100
    expected_asfte.loc[mask_s] = df.loc[mask_s, "Default Hours"] * 24 / 1950 * 100
    expected_asfte.loc[mask_m] = df.loc[mask_m, "Default Hours"] * 12 / 1950 * 100

    # Compare with tolerance 
    asfte_present = df[asfte_col].notna()
    valid_mask = mask_vo_attorney | mask_b | mask_s | mask_m
    bad = asfte_present & valid_mask & ~np.isclose(df[asfte_col], expected_asfte, atol=0.01, rtol=0)

    flag(
        bad,
        "AS FTE % mismatch",
        df,
        highlight_dict,
        [asfte_col, "Default Hours", "Pay Group", "Location"]
    )


    # ---------- export -----------------------------------------------------------
    audit = df[df["Errors"] != ""].copy()
    audit["Errors"] = audit["Errors"].str.rstrip("; ")


    with pd.ExcelWriter(dst, engine="openpyxl") as xls:
        audit.to_excel(xls, sheet_name="Needs Audit", index=False)
        df.to_excel(xls, sheet_name="Original+Flags", index=False)

        # ----- apply cell highlighting ------------------------------------------
        red_fill = PatternFill(start_color="FFFFCCCC", end_color="FFFFCCCC", fill_type="solid")

        # --- Original+Flags sheet -------------------------------------------------
        ws_orig = xls.sheets["Original+Flags"]
        for col_name, rows in highlight_dict.items():
            if not rows:
                continue
            col_idx = df.columns.get_loc(col_name) + 1   # 1‑based
            col_letter = get_column_letter(col_idx)
            for r in rows:
                ws_orig[f"{col_letter}{r + 2}"].fill = red_fill   # +2 → header row offset

        # --- Needs Audit sheet ----------------------------------------------------
        ws_audit = xls.sheets["Needs Audit"]
        # map original DataFrame index → row number in audit sheet
        idx_to_row = {idx: pos + 2 for pos, idx in enumerate(audit.index)}  # header offset
        for col_name, rows in highlight_dict.items():
            if not rows or col_name not in audit.columns:
                continue
            col_idx = audit.columns.get_loc(col_name) + 1
            col_letter = get_column_letter(col_idx)
            for idx in rows:
                if idx in idx_to_row:
                    ws_audit[f"{col_letter}{idx_to_row[idx]}"].fill = red_fill

        # ----- red text for Errors column ----------------------------------------
        err_font = Font(color="FF0000")

        def color_errors(ws, df_source):
            err_col_idx = df_source.columns.get_loc("Errors") + 1
            err_letter = get_column_letter(err_col_idx)
            for row_idx, err_val in enumerate(df_source["Errors"], start=2):  # +2 for header
                if err_val:
                    ws[f"{err_letter}{row_idx}"].font = err_font

        color_errors(ws_orig, df)
        color_errors(ws_audit, audit)

    print(f"{len(audit)} rows need audit → {dst}")
    return dst


# CLI shim
if __name__ == "__main__":
    import sys
    if len(sys.argv) < 2:
        print("Usage: python payroll_autoaudit.py <payroll.xlsx>")
        sys.exit(1)
    main(Path(sys.argv[1]))